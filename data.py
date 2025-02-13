"""
data.py
"""

from openpyxl.reader.excel import load_workbook

class WebData:

    def __init__(self):
        self.url = "https://opensource-demo.orangehrmlive.com/web/index.php/auth/login"
        self.dashboardURL = "https://opensource-demo.orangehrmlive.com/web/index.php/dashboard/index"
        self.fileName ="Data/LoginCredentials.xlsx"
        self.sheetName = "Sheet1"
        self.workbook = load_workbook(self.fileName)
        self.sheet = self.workbook[self.sheetName]


    def rowCount(self):
        """
        Method to return maximum number of rows filled in the Sheet 1 of Excel
        Returns: int

        """
        return self.sheet.max_row

    def readData(self,row,column):
        """
        Method to return the data in the particular cell in the Excel Sheet 1

        """
        return self.sheet.cell(row, column).value

    def writeData(self, row, column, data):
        self.sheet.cell(row, column).value = data
        self.workbook.save(self.fileName)




